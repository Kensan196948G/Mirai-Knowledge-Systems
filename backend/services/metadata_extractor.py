"""
ファイルメタデータ抽出サービス

PDF、Word、Excelなどのファイルからテキストとメタデータを抽出
"""

import io
import logging
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)

# オプション依存関係の遅延インポート
try:
    from PyPDF2 import PdfReader
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False
    logger.debug("PyPDF2未インストール。pip install PyPDF2 でインストールしてください")

try:
    from docx import Document
    PYTHON_DOCX_AVAILABLE = True
except ImportError:
    PYTHON_DOCX_AVAILABLE = False
    logger.debug("python-docx未インストール。pip install python-docx でインストールしてください")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.debug("openpyxl未インストール。pip install openpyxl でインストールしてください")


class MetadataExtractor:
    """ファイルメタデータ抽出器"""

    def __init__(self, metadata_mapping: Optional[Dict[str, str]] = None):
        """
        初期化

        Args:
            metadata_mapping: フィールドマッピングルール
                例: {"Title": "title", "Category": "category"}
        """
        self.metadata_mapping = metadata_mapping or {}

    def extract(
        self,
        file_content: bytes,
        file_name: str,
        sharepoint_metadata: Dict
    ) -> Dict[str, Any]:
        """
        ファイルからメタデータを抽出

        Args:
            file_content: ファイル内容（bytes）
            file_name: ファイル名
            sharepoint_metadata: SharePointメタデータ

        Returns:
            ナレッジレコード用のメタデータ
        """
        file_ext = self._get_file_extension(file_name).lower()

        # ファイル形式別の処理
        if file_ext == ".pdf":
            text_content = self._extract_from_pdf(file_content)
        elif file_ext in [".docx", ".doc"]:
            text_content = self._extract_from_word(file_content)
        elif file_ext in [".xlsx", ".xls"]:
            text_content = self._extract_from_excel(file_content)
        elif file_ext in [".txt", ".md"]:
            text_content = self._extract_from_text(file_content)
        else:
            logger.warning(f"未対応のファイル形式: {file_ext}")
            text_content = f"[{file_ext}ファイル: テキスト抽出未対応]"

        # SharePointメタデータからナレッジフィールドにマッピング
        metadata = self._map_sharepoint_metadata(sharepoint_metadata)

        # ファイル名からタイトル生成（SharePointメタデータにない場合）
        if not metadata.get("title"):
            metadata["title"] = file_name

        # サマリー生成（テキストの最初の500文字）
        if not metadata.get("summary"):
            metadata["summary"] = self._generate_summary(text_content, max_length=500)

        # コンテンツ設定
        metadata["content"] = text_content

        # タグにMS365同期タグを追加
        tags = metadata.get("tags", [])
        if not isinstance(tags, list):
            tags = [tags] if tags else []
        if "MS365同期" not in tags:
            tags.append("MS365同期")
        metadata["tags"] = tags

        return metadata

    def _extract_from_pdf(self, file_content: bytes) -> str:
        """
        PDFからテキストを抽出

        Args:
            file_content: PDFファイル内容

        Returns:
            抽出されたテキスト
        """
        if not PYPDF2_AVAILABLE:
            return "[PDF: PyPDF2未インストールのためテキスト抽出不可]"

        try:
            pdf_file = io.BytesIO(file_content)
            reader = PdfReader(pdf_file)

            text_parts = []
            for page_num, page in enumerate(reader.pages, 1):
                try:
                    text = page.extract_text()
                    if text.strip():
                        text_parts.append(f"--- ページ {page_num} ---\n{text}")
                except Exception as e:
                    logger.warning(f"PDF ページ{page_num}のテキスト抽出エラー: {e}")

            return "\n\n".join(text_parts) if text_parts else "[PDFからテキスト抽出できませんでした]"

        except Exception as e:
            logger.error(f"PDF抽出エラー: {e}")
            return f"[PDFテキスト抽出エラー: {e}]"

    def _extract_from_word(self, file_content: bytes) -> str:
        """
        Wordファイル（.docx）からテキストを抽出

        Args:
            file_content: Wordファイル内容

        Returns:
            抽出されたテキスト
        """
        if not PYTHON_DOCX_AVAILABLE:
            return "[Word: python-docx未インストールのためテキスト抽出不可]"

        try:
            docx_file = io.BytesIO(file_content)
            doc = Document(docx_file)

            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            return "\n\n".join(paragraphs) if paragraphs else "[Wordファイルからテキスト抽出できませんでした]"

        except Exception as e:
            logger.error(f"Word抽出エラー: {e}")
            return f"[Wordテキスト抽出エラー: {e}]"

    def _extract_from_excel(self, file_content: bytes) -> str:
        """
        Excelファイル（.xlsx）からテキストを抽出

        Args:
            file_content: Excelファイル内容

        Returns:
            抽出されたテキスト（最初のシートの内容）
        """
        if not OPENPYXL_AVAILABLE:
            return "[Excel: openpyxl未インストールのためテキスト抽出不可]"

        try:
            xlsx_file = io.BytesIO(file_content)
            workbook = load_workbook(xlsx_file, read_only=True, data_only=True)

            # 最初のシートのみ処理
            sheet = workbook.active
            sheet_name = sheet.title

            rows = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_idx > 100:  # 最大100行まで
                    rows.append("[...（以降の行は省略）...]")
                    break

                # None以外のセルを結合
                cells = [str(cell) if cell is not None else "" for cell in row]
                row_text = " | ".join(cells)
                if row_text.strip():
                    rows.append(row_text)

            text = f"=== {sheet_name} ===\n" + "\n".join(rows)
            return text if rows else "[Excelファイルからテキスト抽出できませんでした]"

        except Exception as e:
            logger.error(f"Excel抽出エラー: {e}")
            return f"[Excelテキスト抽出エラー: {e}]"

    def _extract_from_text(self, file_content: bytes) -> str:
        """
        テキストファイルから内容を抽出

        Args:
            file_content: テキストファイル内容

        Returns:
            テキスト内容
        """
        # エンコーディング自動検出
        encodings = ["utf-8", "utf-8-sig", "shift_jis", "cp932", "euc-jp", "iso-2022-jp"]

        for encoding in encodings:
            try:
                return file_content.decode(encoding)
            except UnicodeDecodeError:
                continue

        # すべて失敗した場合
        logger.warning("テキストファイルのエンコーディング自動検出に失敗")
        return file_content.decode("utf-8", errors="replace")

    def _map_sharepoint_metadata(self, sp_metadata: Dict) -> Dict[str, Any]:
        """
        SharePointメタデータをナレッジフィールドにマッピング

        Args:
            sp_metadata: SharePointメタデータ

        Returns:
            マッピングされたナレッジメタデータ
        """
        mapped = {}

        # カスタムマッピングルール適用
        for sp_field, knowledge_field in self.metadata_mapping.items():
            if sp_field in sp_metadata:
                mapped[knowledge_field] = sp_metadata[sp_field]

        # デフォルトマッピング
        default_mappings = {
            "name": "title",
            "createdBy.user.displayName": "owner",
            "description": "summary",
        }

        for sp_path, knowledge_field in default_mappings.items():
            if knowledge_field not in mapped:
                value = self._get_nested_value(sp_metadata, sp_path)
                if value:
                    mapped[knowledge_field] = value

        return mapped

    def _get_nested_value(self, data: Dict, path: str) -> Any:
        """
        ネストされた辞書から値を取得

        Args:
            data: 辞書
            path: ドットで区切られたパス（例: "createdBy.user.displayName"）

        Returns:
            取得された値（存在しない場合はNone）
        """
        keys = path.split(".")
        current = data

        for key in keys:
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                return None

        return current

    def _get_file_extension(self, file_name: str) -> str:
        """
        ファイル拡張子を取得

        Args:
            file_name: ファイル名

        Returns:
            拡張子（ドット付き、小文字）
        """
        if "." in file_name:
            return "." + file_name.rsplit(".", 1)[1].lower()
        return ""

    def _generate_summary(self, text: str, max_length: int = 500) -> str:
        """
        テキストからサマリーを生成

        Args:
            text: 元のテキスト
            max_length: サマリーの最大長

        Returns:
            サマリー
        """
        # 改行・空白を正規化
        normalized = " ".join(text.split())

        if len(normalized) <= max_length:
            return normalized

        # 最大長で切り詰め（文の途中で切れないように）
        truncated = normalized[:max_length]
        last_period = truncated.rfind("。")
        last_space = truncated.rfind(" ")

        cut_point = max(last_period, last_space)
        if cut_point > max_length * 0.7:  # 70%以上の位置で見つかった場合
            return truncated[:cut_point + 1]

        return truncated + "..."
